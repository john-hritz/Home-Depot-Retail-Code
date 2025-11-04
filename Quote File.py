import os
import pandas as pd
import glob
import re
import datetime
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import date
import matplotlib.pyplot as plt
import shutil

# Get the script's directory
base_path = os.path.dirname(os.path.abspath(__file__))
folder_path = os.path.join(base_path, "Quotes")

def extract_store_date(file_name):
    match = re.search(r'Team Territory Store List (\d{1,2})\.(\d{1,2})\.(\d{4})\.xlsx', file_name)
    if match:
        month, day, year = map(int, match.groups())
        return datetime.date(year, month, day)
    return None

store_files = glob.glob(os.path.join(folder_path, "Team Territory Store List *.xlsx"))
valid_store_files = [f for f in store_files if extract_store_date(f) is not None]
if not valid_store_files:
    raise FileNotFoundError("No valid store listing file found.")
latest_store_file = max(valid_store_files, key=extract_store_date)
store_listing_path = latest_store_file

quote_files = glob.glob(os.path.join(folder_path, "LevitonMfgCoInc-*.xlsx"))
dfs = []

for file in quote_files:
    try:
        df = pd.read_excel(file, sheet_name="Open Quotes")
        dfs.append(df)
    except Exception as e:
        print(f"Error reading {file}: {e}")

if dfs:
    merged_df = pd.concat(dfs, ignore_index=True)
    store_df = pd.read_excel(store_listing_path, sheet_name="Store Listing")
    merged_df = merged_df.merge(store_df[['STORE #', 'DSR']], left_on='StoreNbr', right_on='STORE #', how='left').drop('STORE #', axis=1)
    merged_df['Quote Total'] = pd.to_numeric(merged_df['Quote Total'], errors='coerce')
    df_above = merged_df[merged_df['Quote Total'] > 15000]

    output_file = f"merged_quotes_{date.today().strftime('%Y-%m-%d')}.xlsx"
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_above.to_excel(writer, sheet_name='Above 15000', index=False)

        pivot_table = df_above.pivot_table(
            index=['DSR', 'Quote Date','StoreNbr','ST CD', 'eSVS Order Nbr'],
            values='Quote Total',
            aggfunc='mean'
        ).reset_index()
        pivot_table.to_excel(writer, sheet_name='Summary', index=False)

        workbook = writer.book
        above_15000_sheet = workbook['Above 15000']
        summary_sheet = workbook['Summary']
        
        # Apply filter to row 1 of Above 15000 sheet
        workbook.active = workbook.sheetnames.index('Above 15000')
        above_15000_sheet.auto_filter.ref = above_15000_sheet.dimensions
        above_15000_sheet.sheet_view.zoomScale = 80

        # # Apply filter to row 1 of Summary sheet
        workbook.active = workbook.sheetnames.index('Summary')
        summary_sheet.auto_filter.ref = summary_sheet.dimensions
        
        # Format 'Quote Total' column as currency with no decimals
        currency_format = NamedStyle(name="currency_format", number_format='"$"#,##0')
        for cell in summary_sheet['D'][1:]:  # Skip header
            cell.style = currency_format

        # Set column widths
        for col in summary_sheet.columns:
            summary_sheet.column_dimensions[col[0].column_letter].width = 20
        for col in above_15000_sheet.columns:
            above_15000_sheet.column_dimensions[col[0].column_letter].width = 20

    summary = df_above.groupby('DSR')['eSVS Order Nbr'].nunique().reset_index()
    avg_quote_total = df_above.groupby('DSR')['Quote Total'].mean().reset_index()

    email_body = "<h2>Summary by DSR:</h2><ul>"
    for index, row in summary.iterrows():
        avg_total = avg_quote_total.loc[index, 'Quote Total']
        email_body += f"<li>{row['DSR']}: {row['eSVS Order Nbr']} unique orders over 15000, Avg Quote Total: ${avg_total:,.0f}</li>"
    email_body += "</ul>"

    plt.figure(figsize=(10, 6))
    plt.bar(summary['DSR'], summary['eSVS Order Nbr'], color='skyblue')
    plt.xlabel('DSR')
    plt.ylabel('Number of Unique Orders Over 15000')
    plt.title('Number of Unique Orders Over 15000 by DSR')
    plt.xticks(rotation=45)
    plt.tight_layout()
    chart_file = f"summary_chart_{date.today().strftime('%Y-%m-%d')}.png"
    plt.savefig(chart_file)
    plt.close()

    img_cid = "chart_image"
    email_body += f'<img src="cid:{img_cid}">'

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = "taylor.patterson@leviton.com"
    mail.Subject = f"Quotes Last 3 Weeks {date.today().strftime('%Y-%m-%d')}"
    mail.HTMLBody = email_body
    mail.Attachments.Add(os.path.abspath(output_file))
    attachment = mail.Attachments.Add(os.path.abspath(chart_file))
    attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", img_cid)
    mail.Send()

    print("Email sent successfully!")

    new_folder = os.path.join(folder_path, date.today().strftime('%Y-%m-%d'))
    os.makedirs(new_folder, exist_ok=True)
    for file in quote_files:
        shutil.move(file, new_folder)
else:
    print("No files to merge.")
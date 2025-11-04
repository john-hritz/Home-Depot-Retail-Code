import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import os
import time
import glob
import fnmatch
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import win32com.client
import re
from datetime import datetime, timedelta
import traceback

# Initialize a list to store step durations
step_times = []

# Initialize counts for email
questions_count = 0
reviews_low_rating_count = 0

# Current date and time for age calculations (dynamic)
current_date = datetime.now()
two_weeks_ago = current_date - timedelta(days=14)

# Step 1: Define and create directories, find input files
start_time = time.time()
BASE_DIR = os.path.join(".", "THD Data Warehouse", "Reviews and Questions")
os.makedirs(BASE_DIR, exist_ok=True)

# Search for CSV files
input_files = glob.glob(os.path.join(BASE_DIR, "*.csv"))
print(f"Found input files: {input_files}")
reviews_file = None
questions_file = None
for file in input_files:
    file_basename = os.path.basename(file).lower()
    print(f"Checking file: {file_basename}")
    if "bazaarvoice_network_reviews" in file_basename or fnmatch.fnmatch(file_basename, "bazaarvoice_marketing*.csv"):
        reviews_file = file
        print(f"Selected reviews file: {file}")
    elif "bazaarvoice_network_questions" in file_basename or fnmatch.fnmatch(file_basename, "bazaarvoice_questions*.csv"):
        questions_file = file
        print(f"Selected questions file: {file}")

# Check for Excel file
excel_file = os.path.join(BASE_DIR, "FG Status Report.xlsx")
excel_file_exists = os.path.exists(excel_file)
print(f"Excel file exists: {excel_file_exists} at {excel_file}")

if not reviews_file and not questions_file and not excel_file_exists:
    print(f"Error: No CSV files containing 'bazaarvoice_network_reviews'/'bazaarvoice_marketing*' or 'bazaarvoice_network_questions'/'bazaarvoice_questions*', and no 'FG Status Report.xlsx' found in {BASE_DIR}")
    exit(1)

output_reviews_file = os.path.join(BASE_DIR, "marketing_reviews_to_respond.parquet")
output_questions_file = os.path.join(BASE_DIR, "questions_without_answer.parquet")
output_excel_file = os.path.join(BASE_DIR, "FG_processed_data_grouped.parquet")
output_questions_with_fg_file = os.path.join(BASE_DIR, "questions_with_fg_data.parquet")
output_reviews_with_fg_file = os.path.join(BASE_DIR, "reviews_with_fg_data.parquet")
output_questions_excel = os.path.join(BASE_DIR, "Home Depot Questions.xlsx")
output_reviews_excel = os.path.join(BASE_DIR, "Home Depot Reviews.xlsx")
step1_time = time.time() - start_time
step_times.append(("Directory Setup and File Detection", step1_time))
print(f"Step 1 - Directory Setup and File Detection: {step1_time:.4f} seconds")

# Function to preprocess CSV DataFrame and write to Parquet
def process_csv_file(input_file, output_file, expected_columns, date_columns, string_columns, schema, filter_reviews=False):
    start_time = time.time()
    try:
        # Read CSV file with specified dtypes and date parsing
        dtype_dict = {"UPC": str, "Overall Rating": str}  # Ensure Overall Rating is read as string initially
        df = pd.read_csv(input_file, dtype=dtype_dict)
        available_date_columns = [col for col in date_columns if col in df.columns]
        if available_date_columns:
            df[available_date_columns] = df[available_date_columns].apply(lambda x: pd.to_datetime(x, errors='coerce'))
        # Convert Overall Rating to integer where applicable
        if "Overall Rating" in df.columns:
            df['Overall Rating'] = pd.to_numeric(df['Overall Rating'], errors='coerce', downcast='integer')
        # Filter out 'homedepot-canada' if filter_reviews is True
        if filter_reviews and "Network Destination (Destination Instance)" in df.columns:
            df = df[df['Network Destination (Destination Instance)'] != 'homedepot-canada']
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found.")
        return False, time.time() - start_time, 0
    except Exception as e:
        print(f"Error reading CSV file '{input_file}': {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
        return False, time.time() - start_time, 0

    print(f"Columns in {os.path.basename(input_file)}:", df.columns.tolist())

    # Check for missing required columns
    missing_columns = [col for col in expected_columns if col not in df.columns]
    if missing_columns:
        print(f"Error: Missing required columns in {os.path.basename(input_file)}: {missing_columns}")
        return False, time.time() - start_time, 0

    # Preprocess string columns: fill NaN with empty strings and clean UPC
    for col in string_columns:
        if col in df.columns and col != "Overall Rating":  # Skip Overall Rating as it's now integer
            df[col] = df[col].fillna("").astype(str).str.replace(r'\.0$', '', regex=True)

    # Ensure UPC is explicitly converted to string if present
    if "UPC" in df.columns:
        df["UPC"] = df["UPC"].fillna("").astype(str).str.replace(r'\.0$', '', regex=True)

    # Ensure datetime columns are in datetime64 format
    for col in available_date_columns:
        df[col] = pd.to_datetime(df[col], errors='coerce')

    # Store row count for validation
    input_row_count = len(df)

    # Write to Parquet
    try:
        table = pa.Table.from_pandas(df, schema=schema, preserve_index=False)
        pq.write_table(table, output_file, compression='snappy')
    except Exception as e:
        print(f"Error writing Parquet file '{output_file}': {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
        return False, time.time() - start_time, 0

    # Validate output
    try:
        if os.path.exists(output_file):
            print(f"✅ Parquet file successfully saved to: {output_file}")
            parquet_df = pq.read_table(output_file).to_pandas()
            if "UPC" in parquet_df.columns:
                print(f"Sample UPC values: {parquet_df['UPC'].head().tolist()}")
            for col in available_date_columns:
                print(f"Sample {col} values: {parquet_df[col].head().tolist()}")
            if "Product Page URL" in parquet_df.columns:
                print(f"Sample Product Page URL values: {parquet_df['Product Page URL'].head().tolist()}")
            if "Overall Rating" in parquet_df.columns:
                print(f"Sample Overall Rating values: {parquet_df['Overall Rating'].head().tolist()}")
                print(f"Unique Overall Rating values: {sorted(parquet_df['Overall Rating'].dropna().unique())}")
                print(f"Overall Rating dtype: {parquet_df['Overall Rating'].dtype}")
        else:
            print(f"❌ Parquet file was not created at: {output_file}")
            return False, time.time() - start_time, 0
    except Exception as e:
        print(f"Error validating Parquet file '{output_file}': {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
        return False, time.time() - start_time, 0

    return True, time.time() - start_time, input_row_count

# Function to calculate column width based on content
def calculate_column_width(series, max_width=55, min_width=9):
    max_length = series.astype(str).str.len().max()
    width = min(max(min_width, max_length * 0.8), max_width)
    return round(width, 2)

# Function to sanitize sheet names
def sanitize_sheet_name(name):
    if not name or name.strip() == "":
        return "No_BU"
    invalid_chars = r'[\\*?:/\[\]]'
    sanitized = re.sub(invalid_chars, '_', name.strip())
    return sanitized[:31]  # Excel sheet names max length is 31

# Step 2: Process Reviews file if found
reviews_success = False
reviews_input_row_count = 0
if reviews_file:
    start_time = time.time()
    reviews_expected_columns = [
        "Network Destination (Destination Instance)", "Product Page URL", "Product ID",
        "Product Name", "UPC", "Review First Moderated Date", "Overall Rating",
        "# Days To Respond", "Review Submission Date", "Responder Portal User Name",
        "Reviewer Display Name", "Review Title", "Review Text", "Response Text",
        "Response Submission Date"
    ]
    reviews_date_columns = ["Review Submission Date", "Review First Moderated Date"]
    reviews_string_columns = [
        col for col in reviews_expected_columns if col not in reviews_date_columns and col != "Overall Rating"
    ]
    reviews_schema = pa.schema([
        ("Network Destination (Destination Instance)", pa.string()),
        ("Product Page URL", pa.string()),
        ("Product ID", pa.string()),
        ("Product Name", pa.string()),
        ("UPC", pa.string()),
        ("Review First Moderated Date", pa.timestamp('ns')),
        ("Overall Rating", pa.int32()),
        ("# Days To Respond", pa.string()),
        ("Review Submission Date", pa.timestamp('ns')),
        ("Responder Portal User Name", pa.string()),
        ("Reviewer Display Name", pa.string()),
        ("Review Title", pa.string()),
        ("Review Text", pa.string()),
        ("Response Text", pa.string()),
        ("Response Submission Date", pa.string())
    ])
    print("\nProcessing Reviews file...")
    reviews_success, reviews_time, reviews_input_row_count = process_csv_file(
        reviews_file, output_reviews_file, reviews_expected_columns,
        reviews_date_columns, reviews_string_columns, reviews_schema, filter_reviews=True
    )
    step_times.append(("Process Reviews File", reviews_time))
    print(f"Step 2 - Process Reviews File: {reviews_time:.4f} seconds")

# Step 3: Process Questions file if found
questions_success = False
questions_input_row_count = 0
if questions_file:
    start_time = time.time()
    questions_expected_columns = [
        "Network Destination (Destination Instance)", "Product ID", "Product Page URL", "UPC",
        "Product Name", "Asker Display Name", "Question Submission Date", "Question Title",
        "Question Text", "Answer Submission Date", "Answer Text", "# Days To Answer",
        "Marked \"Can't Answer\""
    ]
    questions_date_columns = ["Question Submission Date"]
    questions_string_columns = [
        col for col in questions_expected_columns if col not in questions_date_columns
    ]
    questions_schema = pa.schema([
        ("Network Destination (Destination Instance)", pa.string()),
        ("Product ID", pa.string()),
        ("Product Page URL", pa.string()),
        ("UPC", pa.string()),
        ("Product Name", pa.string()),
        ("Asker Display Name", pa.string()),
        ("Question Submission Date", pa.timestamp('ns')),
        ("Question Title", pa.string()),
        ("Question Text", pa.string()),
        ("Answer Submission Date", pa.string()),
        ("Answer Text", pa.string()),
        ("# Days To Answer", pa.string()),
        ("Marked \"Can't Answer\"", pa.string())
    ])
    print("\nProcessing Questions file...")
    questions_success, questions_time, questions_input_row_count = process_csv_file(
        questions_file, output_questions_file, questions_expected_columns,
        questions_date_columns, questions_string_columns, questions_schema
    )
    step_times.append(("Process Questions File", questions_time))
    print(f"Step 3 - Process Questions File: {questions_time:.4f} seconds")

# Step 4: Process Excel file if found
excel_success = False
if excel_file_exists:
    start_time = time.time()
    print("\nProcessing Excel file...")
    try:
        dtype_dict = {"UPC": str, "M P G": str, "IPG": str, "BU": str, "Basic": str, "Dash": str}
        df = pd.read_excel(excel_file, sheet_name="Sheet1", engine='openpyxl', dtype=dtype_dict)
    except FileNotFoundError:
        print(f"Error: Input file '{excel_file}' not found.")
        step_times.append(("Process Excel File", time.time() - start_time))
        print(f"Step 4 - Process Excel File: {(time.time() - start_time):.4f} seconds")
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
        step_times.append(("Process Excel File", time.time() - start_time))
        print(f"Step 4 - Process Excel File: {(time.time() - start_time):.4f} seconds")
    else:
        print("Columns in Excel file:", df.columns.tolist())
        required_columns = ["UPC", "M P G", "IPG", "BU", "Basic", "Dash"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"Error: Missing columns in Excel file: {missing_columns}")
            step_times.append(("Process Excel File", time.time() - start_time))
            print(f"Step 4 - Process Excel File: {(time.time() - start_time):.4f} seconds")
        else:
            df = df[required_columns]
            for col in required_columns:
                df[col] = df[col].fillna("").astype(str).str.replace(r'\.0$', '', regex=True)
            try:
                grouped_df = df.groupby('UPC').agg({
                    'M P G': 'first',
                    'IPG': 'first',
                    'BU': 'first',
                    'Basic': 'first',
                    'Dash': 'first'
                }).reset_index()
                grouped_df['M P G'] = grouped_df['M P G'].astype(str)
                grouped_df['IPG'] = grouped_df['IPG'].astype(str)
                grouped_df['MPG_IPG'] = grouped_df['M P G'] + '-' + grouped_df['IPG']
                grouped_df['MPG_IPG'] = grouped_df['MPG_IPG'].replace(['-', ''], 'Unknown')
                grouped_df['Basic_Dash'] = grouped_df['Basic'] + '-' + grouped_df['Dash']
            except Exception as e:
                print(f"Error during Excel data processing: {str(e)}")
                print(f"Exception details: {traceback.format_exc()}")
                step_times.append(("Process Excel File", time.time() - start_time))
                print(f"Step 4 - Process Excel File: {(time.time() - start_time):.4f} seconds")
            else:
                try:
                    schema = pa.schema([
                        ('UPC', pa.string()),
                        ('M P G', pa.string()),
                        ('IPG', pa.string()),
                        ('BU', pa.string()),
                        ('MPG_IPG', pa.string()),
                        ('Basic', pa.string()),
                        ('Dash', pa.string()),
                        ('Basic_Dash', pa.string())
                    ])
                    table = pa.Table.from_pandas(grouped_df, schema=schema, preserve_index=False)
                    pq.write_table(table, output_excel_file, compression='snappy')
                except Exception as e:
                    print(f"Error writing Excel Parquet file: {str(e)}")
                    print(f"Exception details: {traceback.format_exc()}")
                    step_times.append(("Process Excel File", time.time() - start_time))
                    print(f"Step 4 - Process Excel File: {(time.time() - start_time):.4f} seconds")
                else:
                    try:
                        if os.path.exists(output_excel_file):
                            print(f"✅ Parquet file successfully saved to: {output_excel_file}")
                            parquet_df = pq.read_table(output_excel_file).to_pandas()
                            print(f"Sample UPC values: {parquet_df['UPC'].head().tolist()}")
                            print(f"Sample M P G values: {parquet_df['M P G'].head().tolist()}")
                            print(f"Sample IPG values: {parquet_df['IPG'].head().tolist()}")
                            print(f"Sample BU values: {parquet_df['BU'].head().tolist()}")
                            print(f"Sample MPG_IPG values: {parquet_df['MPG_IPG'].head().tolist()}")
                            print(f"Sample Basic_Dash values: {parquet_df['Basic_Dash'].head().tolist()}")
                            unique_upcs = grouped_df['UPC'].nunique()
                            total_rows = len(grouped_df)
                            if unique_upcs == total_rows:
                                print("✅ Each UPC has exactly one row.")
                            else:
                                print("❌ Duplicate UPCs found.")
                                print(f"Unique UPCs: {unique_upcs}")
                                print(f"Total Rows: {total_rows}")
                                duplicates = grouped_df['UPC'].value_counts()[grouped_df['UPC'].value_counts() > 1]
                                print("Duplicate UPCs:\n", duplicates)
                        else:
                            print(f"❌ Parquet file was not created at: {output_excel_file}")
                            excel_success = False
                        excel_success = True
                    except Exception as e:
                        print(f"Error validating Excel Parquet file: {str(e)}")
                        print(f"Exception details: {traceback.format_exc()}")
                    step_times.append(("Process Excel File", time.time() - start_time))
                    print(f"Step 4 - Process Excel File: {(time.time() - start_time):.4f} seconds")

# Step 5: Create questions_with_fg_data.parquet
questions_with_fg_success = False
if questions_success and excel_success:
    start_time = time.time()
    print("\nCreating questions_with_fg_data.parquet...")
    try:
        questions_df = pq.read_table(output_questions_file).to_pandas()
        fg_df = pq.read_table(output_excel_file).to_pandas()
        questions_df['UPC_11'] = questions_df['UPC'].str[:11]
        fg_df['UPC_11'] = fg_df['UPC'].str[:11]
        merged_df = questions_df.merge(
            fg_df[['UPC_11', 'M P G', 'IPG', 'BU', 'MPG_IPG', 'Basic_Dash']],
            on='UPC_11',
            how='left'
        )
        merged_df = merged_df.rename(columns={'M P G': 'MPG'})
        merged_df = merged_df.drop(columns=['UPC_11'])
        date_columns = ['Question Submission Date']
        string_columns = [col for col in merged_df.columns if col not in date_columns]
        for col in string_columns:
            merged_df[col] = merged_df[col].fillna("").astype(str).str.replace(r'\.0$', '', regex=True)
        schema = pa.schema([
            ("Network Destination (Destination Instance)", pa.string()),
            ("Product ID", pa.string()),
            ("Product Page URL", pa.string()),
            ("UPC", pa.string()),
            ("Product Name", pa.string()),
            ("Asker Display Name", pa.string()),
            ("Question Submission Date", pa.timestamp('ns')),
            ("Question Title", pa.string()),
            ("Question Text", pa.string()),
            ("Answer Submission Date", pa.string()),
            ("Answer Text", pa.string()),
            ("# Days To Answer", pa.string()),
            ("Marked \"Can't Answer\"", pa.string()),
            ("MPG", pa.string()),
            ("IPG", pa.string()),
            ("BU", pa.string()),
            ("MPG_IPG", pa.string()),
            ("Basic_Dash", pa.string())
        ])
        table = pa.Table.from_pandas(merged_df, schema=schema, preserve_index=False)
        pq.write_table(table, output_questions_with_fg_file, compression='snappy')
        if os.path.exists(output_questions_with_fg_file):
            print(f"✅ Parquet file successfully saved to: {output_questions_with_fg_file}")
            parquet_df = pq.read_table(output_questions_with_fg_file).to_pandas()
            print(f"Total rows in questions_with_fg_data.parquet: {len(parquet_df)}")
            if len(parquet_df) == questions_input_row_count:
                print("✅ Row count matches input questions file.")
            else:
                print(f"⚠️ Row count differs from input questions file ({questions_input_row_count} rows).")
            print(f"Columns in questions_with_fg_data.parquet: {parquet_df.columns.tolist()}")
            print(f"Sample UPC values: {parquet_df['UPC'].head().tolist()}")
            print(f"Sample Question Submission Date values: {parquet_df['Question Submission Date'].head().tolist()}")
            print(f"Sample MPG values: {parquet_df['MPG'].head().tolist()}")
            print(f"Sample IPG values: {parquet_df['IPG'].head().tolist()}")
            print(f"Sample BU values: {parquet_df['BU'].head().tolist()}")
            print(f"Sample MPG_IPG values: {parquet_df['MPG_IPG'].head().tolist()}")
            print(f"Sample Basic_Dash values: {parquet_df['Basic_Dash'].head().tolist()}")
            questions_with_fg_success = True
        else:
            print(f"❌ Parquet file was not created at: {output_questions_with_fg_file}")
    except Exception as e:
        print(f"Error creating questions_with_fg_data.parquet: {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
    step_times.append(("Create Questions with FG Data", time.time() - start_time))
    print(f"Step 5 - Create Questions with FG Data: {(time.time() - start_time):.4f} seconds")

# Step 6: Create Home Depot Questions Excel file
questions_excel_success = False
if questions_with_fg_success:
    start_time = time.time()
    print("\nCreating Home Depot Questions Excel file...")
    try:
        questions_df = pq.read_table(output_questions_with_fg_file).to_pandas()
        required_columns = [
            "Basic_Dash", "BU", "MPG", "IPG", "Product ID", "Product Page URL", "UPC",
            "Question Submission Date", "Question Title", "Question Text", "Answer Text",
            "Marked \"Can't Answer\""
        ]
        missing_columns = [col for col in required_columns if col not in questions_df.columns]
        if missing_columns:
            print(f"Error: Missing columns in questions_with_fg_data.parquet: {missing_columns}")
        else:
            questions_df = questions_df[required_columns]
            questions_df['BU'] = questions_df['BU'].fillna("").astype(str)
            questions_df = questions_df.sort_values(['Basic_Dash', 'Question Submission Date'], ascending=[False, False])
            questions_count = len(questions_df)  # Calculate total questions count
            # Calculate row counts and oldest dates per BU for email
            questions_bu_stats = questions_df.groupby('BU').agg({
                'Question Submission Date': ['count', 'min']
            }).reset_index()
            questions_bu_stats.columns = ['BU', 'count', 'oldest_date']
            questions_bu_stats['old_count'] = questions_df[questions_df['Question Submission Date'] < two_weeks_ago].groupby('BU').size().reindex(questions_bu_stats['BU'], fill_value=0).values
            questions_bu_stats['oldest_date_str'] = questions_bu_stats['oldest_date'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notnull(x) else 'No Date')
            questions_bu_text = "\n".join([
                f"  - {row['BU'] if row['BU'] else 'No BU'}: {row['count']} rows ({row['old_count']} older than 2 weeks, oldest: {row['oldest_date_str']})"
                for _, row in questions_bu_stats.iterrows()
            ])
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            bu_list = sorted(questions_df['BU'].unique(), reverse=False)  # Sort BU tabs A to Z
            for bu in bu_list:
                sheet_name = sanitize_sheet_name(bu)
                sheet = wb.create_sheet(title=sheet_name)
                bu_df = questions_df[questions_df['BU'] == bu]
                headers = required_columns
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)  # Apply bold to header row
                for row_idx, row in enumerate(bu_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        sheet.cell(row=row_idx, column=col_idx).value = value
                sheet.freeze_panes = 'A2'
                sheet.sheet_view.zoomScale = 80
                # Apply autofilter to header row
                last_col_letter = get_column_letter(len(headers))
                sheet.auto_filter.ref = f"A1:{last_col_letter}{sheet.max_row}"
                for col_idx, column in enumerate(headers, 1):
                    col_letter = get_column_letter(col_idx)
                    if column == "Answer Text":
                        sheet.column_dimensions[col_letter].width = 55
                    elif column == "Question Submission Date":
                        sheet.column_dimensions[col_letter].width = 19
                    else:
                        width = calculate_column_width(bu_df[column], max_width=55, min_width=9)
                        sheet.column_dimensions[col_letter].width = width
                    for row_idx in range(1, sheet.max_row + 1):
                        sheet[f"{col_letter}{row_idx}"].alignment = openpyxl.styles.Alignment(wrap_text=True)
            wb.save(output_questions_excel)
            if os.path.exists(output_questions_excel):
                print(f"✅ Excel file successfully saved to: {output_questions_excel}")
                questions_excel_success = True
            else:
                print(f"❌ Excel file was not created at: {output_questions_excel}")
    except Exception as e:
        print(f"Error creating Home Depot Questions Excel file: {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
    step_times.append(("Create Home Depot Questions Excel", time.time() - start_time))
    print(f"Step 6 - Create Home Depot Questions Excel: {(time.time() - start_time):.4f} seconds")

# Step 7: Create reviews_with_fg_data.parquet
reviews_with_fg_success = False
if reviews_success and excel_success:
    start_time = time.time()
    print("\nCreating reviews_with_fg_data.parquet...")
    try:
        reviews_df = pq.read_table(output_reviews_file).to_pandas()
        fg_df = pq.read_table(output_excel_file).to_pandas()
        reviews_df['UPC_11'] = reviews_df['UPC'].str[:11]
        fg_df['UPC_11'] = fg_df['UPC'].str[:11]
        merged_df = reviews_df.merge(
            fg_df[['UPC_11', 'M P G', 'IPG', 'BU', 'MPG_IPG', 'Basic_Dash']],
            on='UPC_11',
            how='left'
        )
        merged_df = merged_df.rename(columns={'M P G': 'MPG'})
        merged_df = merged_df.drop(columns=['UPC_11'])
        # Ensure Overall Rating is numeric
        merged_df['Overall Rating'] = pd.to_numeric(merged_df['Overall Rating'], errors='coerce', downcast='integer')
        # Ensure Review First Moderated Date is datetime (since Review Submission Date may be missing)
        if 'Review First Moderated Date' in merged_df.columns:
            merged_df['Review First Moderated Date'] = pd.to_datetime(merged_df['Review First Moderated Date'], errors='coerce')
        if 'Review Submission Date' in merged_df.columns:
            merged_df['Review Submission Date'] = pd.to_datetime(merged_df['Review Submission Date'], errors='coerce')
        # Clean string columns
        date_columns = ['Review Submission Date', 'Review First Moderated Date']
        string_columns = [col for col in merged_df.columns if col not in date_columns and col != 'Overall Rating']
        for col in string_columns:
            merged_df[col] = merged_df[col].fillna("").astype(str).str.replace(r'\.0$', '', regex=True)
        schema = pa.schema([
            ("Network Destination (Destination Instance)", pa.string()),
            ("Product Page URL", pa.string()),
            ("Product ID", pa.string()),
            ("Product Name", pa.string()),
            ("UPC", pa.string()),
            ("Review First Moderated Date", pa.timestamp('ns')),
            ("Overall Rating", pa.int32()),
            ("# Days To Respond", pa.string()),
            ("Review Submission Date", pa.timestamp('ns')),
            ("Responder Portal User Name", pa.string()),
            ("Reviewer Display Name", pa.string()),
            ("Review Title", pa.string()),
            ("Review Text", pa.string()),
            ("Response Text", pa.string()),
            ("Response Submission Date", pa.string()),
            ("MPG", pa.string()),
            ("IPG", pa.string()),
            ("BU", pa.string()),
            ("MPG_IPG", pa.string()),
            ("Basic_Dash", pa.string())
        ])
        table = pa.Table.from_pandas(merged_df, schema=schema, preserve_index=False)
        pq.write_table(table, output_reviews_with_fg_file, compression='snappy')
        if os.path.exists(output_reviews_with_fg_file):
            print(f"✅ Parquet file successfully saved to: {output_reviews_with_fg_file}")
            parquet_df = pq.read_table(output_reviews_with_fg_file).to_pandas()
            print(f"Total rows in reviews_with_fg_data.parquet: {len(parquet_df)}")
            if len(parquet_df) == reviews_input_row_count:
                print("✅ Row count matches input reviews file.")
            else:
                print(f"⚠️ Row count differs from input reviews file ({reviews_input_row_count} rows).")
            print(f"Columns in reviews_with_fg_data.parquet: {parquet_df.columns.tolist()}")
            print(f"Sample UPC values: {parquet_df['UPC'].head().tolist()}")
            if 'Review Submission Date' in parquet_df.columns:
                print(f"Sample Review Submission Date values: {parquet_df['Review Submission Date'].head().tolist()}")
                null_dates = parquet_df['Review Submission Date'].isna().sum()
                print(f"Null Review Submission Date values: {null_dates}")
            if 'Review First Moderated Date' in parquet_df.columns:
                print(f"Sample Review First Moderated Date values: {parquet_df['Review First Moderated Date'].head().tolist()}")
                null_dates = parquet_df['Review First Moderated Date'].isna().sum()
                print(f"Null Review First Moderated Date values: {null_dates}")
            print(f"Sample MPG values: {parquet_df['MPG'].head().tolist()}")
            print(f"Sample IPG values: {parquet_df['IPG'].head().tolist()}")
            print(f"Sample BU values: {parquet_df['BU'].head().tolist()}")
            print(f"Sample MPG_IPG values: {parquet_df['MPG_IPG'].head().tolist()}")
            print(f"Sample Basic_Dash values: {parquet_df['Basic_Dash'].head().tolist()}")
            print(f"Unique Overall Rating values: {sorted(parquet_df['Overall Rating'].dropna().unique())}")
            print(f"Overall Rating dtype: {parquet_df['Overall Rating'].dtype}")
            reviews_with_fg_success = True
        else:
            print(f"❌ Parquet file was not created at: {output_reviews_with_fg_file}")
    except Exception as e:
        print(f"Error creating reviews_with_fg_data.parquet: {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
    step_times.append(("Create Reviews with FG Data", time.time() - start_time))
    print(f"Step 7 - Create Reviews with FG Data: {(time.time() - start_time):.4f} seconds")

# Step 8: Create Home Depot Reviews Excel file and apply filter
reviews_excel_success = False
if reviews_with_fg_success:
    start_time = time.time()
    print("\nCreating Home Depot Reviews Excel file...")
    try:
        reviews_df = pq.read_table(output_reviews_with_fg_file).to_pandas()
        print(f"Columns in reviews_with_fg_data.parquet: {reviews_df.columns.tolist()}")
        required_columns = [
            "Basic_Dash", "BU", "MPG", "IPG", "Product ID", "Product Name", "Product Page URL",
            "UPC", "Review First Moderated Date", "Overall Rating", "Review Text", "Response Text"
        ]
        missing_columns = [col for col in required_columns if col not in reviews_df.columns]
        if missing_columns:
            print(f"Error: Missing columns in reviews_with_fg_data.parquet: {missing_columns}")
        else:
            reviews_df = reviews_df[required_columns]
            reviews_df['BU'] = reviews_df['BU'].fillna("").astype(str)
            reviews_df['Overall Rating'] = pd.to_numeric(reviews_df['Overall Rating'], errors='coerce', downcast='integer')
            # Ensure Review First Moderated Date is datetime
            reviews_df['Review First Moderated Date'] = pd.to_datetime(reviews_df['Review First Moderated Date'], errors='coerce')
            print(f"Unique Overall Rating values before Excel: {sorted(reviews_df['Overall Rating'].dropna().unique())}")
            print(f"Overall Rating dtype before Excel: {reviews_df['Overall Rating'].dtype}")
            reviews_low_rating_count = len(reviews_df[reviews_df['Overall Rating'].isin([1, 2, 3])])  # Calculate low rating count
            # Use Review First Moderated Date for stats, as Review Submission Date is not in output
            date_col = 'Review First Moderated Date'
            reviews_bu_stats = reviews_df.groupby('BU').agg({
                date_col: ['count', 'min'],
                'Overall Rating': lambda x: sum(x.isin([1, 2, 3]))
            }).reset_index()
            reviews_bu_stats.columns = ['BU', 'count', 'oldest_date', 'low_rating_count']
            reviews_bu_stats['old_count'] = reviews_df[reviews_df[date_col] < two_weeks_ago].groupby('BU').size().reindex(reviews_bu_stats['BU'], fill_value=0).values
            reviews_bu_stats['oldest_date_str'] = reviews_bu_stats['oldest_date'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notnull(x) else 'No Date')
            reviews_bu_text = "\n".join([
                f"  - {row['BU'] if row['BU'] else 'No BU'}: {row['count']} rows ({row['low_rating_count']} with ratings 1, 2, or 3, {row['old_count']} older than 2 weeks, oldest: {row['oldest_date_str']})"
                for _, row in reviews_bu_stats.iterrows()
            ])
            reviews_df = reviews_df.sort_values(['Basic_Dash', 'Review First Moderated Date'], ascending=[False, False])
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            bu_list = sorted(reviews_df['BU'].unique(), reverse=False)  # Sort BU tabs A to Z
            for bu in bu_list:
                sheet_name = sanitize_sheet_name(bu)
                sheet = wb.create_sheet(title=sheet_name)
                bu_df = reviews_df[reviews_df['BU'] == bu]
                headers = required_columns
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                for row_idx, row in enumerate(bu_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        if headers[col_idx - 1] == "Overall Rating" and value is not None:
                            cell.number_format = '0'
                sheet.freeze_panes = 'A2'
                sheet.sheet_view.zoomScale = 80
                last_col_letter = get_column_letter(len(headers))
                sheet.auto_filter.ref = f"A1:{last_col_letter}{sheet.max_row}"
                for col_idx, column in enumerate(headers, 1):
                    col_letter = get_column_letter(col_idx)
                    if column == "Response Text":
                        sheet.column_dimensions[col_letter].width = 55
                    else:
                        width = calculate_column_width(bu_df[column], max_width=55, min_width=9)
                        sheet.column_dimensions[col_letter].width = width
                    for row_idx in range(1, sheet.max_row + 1):
                        sheet[f"{col_letter}{row_idx}"].alignment = openpyxl.styles.Alignment(wrap_text=True)
            wb.save(output_reviews_excel)
            if os.path.exists(output_reviews_excel):
                print(f"✅ Excel file successfully saved to: {output_reviews_excel}")
                reviews_excel_success = True
                try:
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    workbook = excel.Workbooks.Open(os.path.abspath(output_reviews_excel))
                    for sheet in workbook.Sheets:
                        if sheet.AutoFilterMode:
                            sheet.AutoFilterMode = False
                        sheet.Range("A1").AutoFilter(Field=10, Criteria1=["=1", "=2", "=3"], Operator=7)
                        visible_rows = sum(1 for row in range(2, sheet.UsedRange.Rows.Count + 1) if not sheet.Rows(row).Hidden)
                        print(f"Sheet '{sheet.Name}' has {visible_rows} visible rows after filtering (ratings 1, 2, or 3).")
                    workbook.Save()
                    workbook.Close()
                    excel.Quit()
                    print("✅ Filter applied successfully to show only ratings 1, 2, or 3.")
                except Exception as e:
                    print(f"Error applying filter with win32com.client: {str(e)}")
                    print(f"Exception details: {traceback.format_exc()}")
            else:
                print(f"❌ Excel file was not created at: {output_reviews_excel}")
    except Exception as e:
        print(f"Error creating Home Depot Reviews Excel file: {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
    step_times.append(("Create Home Depot Reviews Excel", time.time() - start_time))
    print(f"Step 8 - Create Home Depot Reviews Excel: {(time.time() - start_time):.4f} seconds")

# Step 9: Send email with Excel files
email_success = False
if questions_excel_success or reviews_excel_success:
    start_time = time.time()
    print("\nSending email with Excel files...")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0 = MailItem
        mail.To = "john.hritz@leviton.com"
        mail.Subject = "Home Depot Questions and Reviews Excel Files"
        email_body = f"""
<html>
<body>
<p>Below are links to Home Depot Questions and Answers to respond to. Please input response directly into the file at the links.  
Files will lock Tuesday at 4 PM EST to allow upload of responses.  New files will be sent out Wedneday morning:</p>
<ul>
"""
        if questions_excel_success:
            email_body += f"""
    <li><a href="https://leviton.sharepoint.com/:x:/r/sites/Retail/_layouts/15/Doc.aspx?sourcedoc=%7B1877F45A-C7C9-458F-AA98-324E7E8491A2%7D&file=Home%20Depot%20Questions.xlsx&action=default&mobileredirect=true">Home Depot Questions Link</a>: Contains {questions_count} total questions, with tabs for each BU.</li>
    <p>Rows per BU tab (with counts older than 2 weeks and oldest date):</p>
    <pre>{questions_bu_text}</pre>
"""
        else:
            email_body += "<li>Home Depot Questions file could not be generated due to an error.</li>"

        if reviews_excel_success:
            email_body += f"""
    <li><a href="https://leviton.sharepoint.com/:x:/r/sites/Retail/_layouts/15/Doc.aspx?sourcedoc=%7BBBB559CC-6DEF-4E1C-ACEA-F447DC33813E%7D&file=Home%20Depot%20Reviews.xlsx&action=default&mobileredirect=true">Home Depot Reviews Link</a>: There are {reviews_low_rating_count} reviews with Overall Rating of 1, 2, or 3. Autofilter set to show only ratings of 1, 2, or 3 by default, but all ratings are included and accessible via the filter dropdown.</li>
    <p>Rows per BU tab (with low rating counts, counts older than 2 weeks, and oldest date based on Review First Moderated Date):</p>
    <pre>{reviews_bu_text}</pre>
"""
        else:
            email_body += "<li>Home Depot Reviews file could not be generated due to an error.</li>"

        email_body += """
</ul>
<p>Please review and let me know if you need further details.</p>
</body>
</html>
"""
        mail.HTMLBody = email_body
        for file in [output_questions_excel, output_reviews_excel]:
            if os.path.exists(file):
                mail.Attachments.Add(os.path.abspath(file))
            else:
                print(f"Error: Attachment file '{file}' not found.")
        mail.Send()
        print("✅ Email sent successfully to john.hritz@leviton.com")
        email_success = True
    except Exception as e:
        print(f"Error sending email: {str(e)}")
        print(f"Exception details: {traceback.format_exc()}")
    step_times.append(("Send Email with Excel Files", time.time() - start_time))
    print(f"Step 9 - Send Email with Excel Files: {(time.time() - start_time):.4f} seconds")

# Step 10: Print execution time summary
total_time = sum(time for _, time in step_times)
print("\n=== Execution Time Summary ===")
for step, duration in step_times:
    print(f"{step}: {duration:.4f} seconds")
print(f"Total Execution Time: {total_time:.4f} seconds")
if reviews_success:
    print(f"Reviews Parquet file saved to: {output_reviews_file}")
if questions_success:
    print(f"Questions Parquet file saved to: {output_questions_file}")
if excel_success:
    print(f"Excel Parquet file saved to: {output_excel_file}")
if questions_with_fg_success:
    print(f"Questions with FG Data Parquet file saved to: {output_questions_with_fg_file}")
if reviews_with_fg_success:
    print(f"Reviews with FG Data Parquet file saved to: {output_reviews_with_fg_file}")
if questions_excel_success:
    print(f"Questions Excel file saved to: {output_questions_excel}")
if reviews_excel_success:
    print(f"Reviews Excel file saved to: {output_reviews_excel}")
if email_success:
    print("Email sent with Questions and Reviews Excel files.")
if not (reviews_success or questions_success or excel_success or questions_with_fg_success or
        reviews_with_fg_success or questions_excel_success or reviews_excel_success):
    print("No files were successfully processed.")
import polars as pl
import os
import psutil
import time
import pandas as pd
import openpyxl
import argparse
import json
import shutil
from datetime import datetime
from typing import Dict, Optional, List
from polars import DataType

# Constants
BASE_DIR = os.path.join(os.getcwd(), 'THD Data Warehouse')
DATA_TABLES_DIR = os.path.join(BASE_DIR, 'Data_Tables')
PARQUET_VERSIONS_DIR = os.path.join(BASE_DIR, 'parquet_versions')
log_file = os.path.join(BASE_DIR, 'process.log')
os.makedirs(DATA_TABLES_DIR, exist_ok=True)
os.makedirs(PARQUET_VERSIONS_DIR, exist_ok=True)

# File configurations: (file prefix, column types, output Parquet name, is_excel, search_columns, sheet_name, start_row)
file_configs = [
    (
        "pythononlinewebsiteanalysis",
        {
            "day": pl.Datetime,
            "week": pl.Utf8,
            "oms id +": pl.Int32,
            "online upc cd +": pl.Utf8,
            "online product interaction conversion rate": pl.Float32,  # Corrected to match Parquet (no trailing space)
            "online product interaction conversion rate ly": pl.Float32,
            "online pip conversion rate +": pl.Float32,
            "online pip conversion rate ly +": pl.Float32,
            "online display avg rating +": pl.Float32,
            "online display count of 1 star reviews +": pl.Float32,
            "online display count of 2 star reviews +": pl.Float32,
            "online display count of 3 star reviews +": pl.Float32,
            "online display count of 4 star reviews +": pl.Float32,
            "online display count of 5 star reviews +": pl.Float32,
            "online non-buyable views +": pl.Float32,
            "online non-buyable views ly +": pl.Float32,
            "online current cost $ +": pl.Float32,
            "online pip visits +": pl.Float32,
            "online pip visits ly +": pl.Float32,
            "online product interaction visits +": pl.Float32,
            "online product interaction visits ly +": pl.Float32,
            "order count TY": pl.Float32,
            "order count LY": pl.Float32,
        },
        "online_website_anaylsis.parquet",
        False,
        ["day", "oms id +"],
        None,
        1
    ),
    (
        "pythononlinesales",
        {
            "day": pl.Datetime,
            "fulfillment channel +": pl.Categorical,
            "fulfillment channel name +": pl.Categorical,
            "icr store +": pl.Categorical,
            "online upc +": pl.Categorical,
            "week": pl.Categorical,
            "oms id +": pl.Int32,
            "online sales $ +": pl.Float32,
            "online sales $ ly +": pl.Float32,
            "online cancel units +": pl.Int32,
            "online cancel units ly +": pl.Int32,
            "online count of orders +": pl.Int32,
            "online count of orders ly +": pl.Int32,
            "online gross demand $ +": pl.Float32,
            "online gross demand $ ly +": pl.Float32,
            "online order units +": pl.Int32,
            "online order units ly +": pl.Int32,
            "online return $ +": pl.Float32,
            "online return $ ly +": pl.Float32,
            "online return units +": pl.Float32,
            "online return units ly +": pl.Int32,
            "online sales $ before returns +": pl.Float32,
            "online sales $ before returns ly +": pl.Float32,
            "online settled units +": pl.Int32,
            "online settled units before returns +": pl.Int32,
            "online settled units before returns ly +": pl.Int32,
            "online settled units ly+": pl.Int32,
        },
        "online_sales.parquet",
        False,
        ["day", "oms id +"],
        None,
        1
    ),
    (
        "pythononlineclassification",
        {
            "oms id +": pl.Int32,
            "online class +": pl.Categorical,
            "online merch dept +": pl.Categorical,
            "online subclass +": pl.Categorical,
            "internet sku name +": pl.Categorical,
            "online merch dept nbr +": pl.Categorical,
            "online classification +": pl.String,
            "online disposition": pl.String,
            "online ship from type +": pl.Categorical,
            "online ship lead days": pl.Categorical,
            "online upc +": pl.Utf8,
        },
        "online_classification.parquet",
        False,
        None,
        None,
        1
    ),
    (
        "BA_VendorContentScorecard",
        {
            "TOTAL SCORE": pl.Float32,
            "# ENRICHMENTS": pl.Int32,
            "OMSID": pl.Int32,
            "CLASS BAND": pl.Categorical,
            "MFG MODEL": pl.String,
            "VENDOR NAME": pl.Categorical,
            "MFG BRAND NAME": pl.Categorical,
            "ALT IMAGES # ": pl.Int64,
            "ALT IMAGES TARGET": pl.Int32,
            "": pl.String,
            "DOCS # ": pl.Int64,
            "_duplicated_0": pl.String,
            "DOCS TARGET": pl.Int32,
            "VIDEOS # ": pl.Int64,
            "_duplicated_1": pl.String,
            "VIDEOS TARGET": pl.Int32,
            "_duplicated_2": pl.String,
            "SALIENT BULLETS # ": pl.Int64,
            "_duplicated_3": pl.String,
            "SALIENT BULLETS TARGET": pl.Int32,
            "AUGMENTED REALITY Y/N": pl.Categorical,
            "_duplicated_4": pl.String,
            "_duplicated_5": pl.String,
            "_duplicated_6": pl.String,
            "_duplicated_7": pl.String,
            "_duplicated_8": pl.String,
            "_duplicated_9": pl.String,
            "AUGMENTED REALITY TARGET": pl.String,
            "_duplicated_10": pl.String,
            "_duplicated_11": pl.String,
            "360 SPIN Y/N": pl.Categorical,
            "_duplicated_12": pl.String,
            "_duplicated_13": pl.String,
            "360 SPIN TARGET": pl.String,
            "REVIEWS # ": pl.Int64,
            "_duplicated_14": pl.String,
            "REViEWS TARGET": pl.Int64,
            "SSKU Y/N": pl.Categorical,
            "_duplicated_15": pl.String,
            "_duplicated_16": pl.String,
            "SSKU TARGET": pl.String,
            "NODE NAME": pl.Categorical,
            "_duplicated_17": pl.String,
            "SUB DEPT": pl.Categorical,
            "_duplicated_18": pl.String,
            "CLASS": pl.Categorical,
            "SUB CLASS": pl.Categorical,
            "PRODUCT NAME 120": pl.String,
        },
        "BA_scorecard.parquet",
        True,
        None,
        "Summary",
        28
    ),
    (
        "onlinestores",
        {
            "icr byo name +": pl.Categorical,
            "icr byo nbr +": pl.Int32,
            "icr buying office +": pl.Categorical,
            "icr city +": pl.Utf8,
            "icr county +": pl.Utf8,
            "icr district  +": pl.Utf8,
            "icr district name +": pl.Categorical,
            "icr district nbr +": pl.Int32,
            "icr division +": pl.Categorical,
            "icr division name +": pl.Categorical,
            "icr division nbr +": pl.Int32,
            "icr latitude +": pl.Float64,
            "icr longitude +": pl.Float64,
            "icr market name +": pl.Categorical,
            "icr market nbr +": pl.Int32,
            "icr store name +": pl.Utf8,
            "icr store nbr +": pl.Utf8,
            "icr postal code +": pl.Utf8,
            "icr region  +": pl.Utf8,
            "icr region name +": pl.Utf8,
            "icr region nbr +": pl.Int32,
            "icr state territory code +": pl.Utf8,
            "icr market +": pl.Utf8,
        },
        "online_stores.parquet",
        False,
        None,
        None,
        1
    ),
    (
        "PythonCalendar_Full",
        {
            "week": pl.Utf8,
        },
        "calendar.parquet",
        False,
        None,
        None,
        1
    ),
    (
        "FG Status Report",
        {
            "External": pl.Utf8,
            "Basic San": pl.Utf8,
            "Wse": pl.Utf8,
            "Plt": pl.Utf8,
            "Dpt": pl.Utf8,
            "Pln Ctr": pl.Utf8,
            "GroupNbr": pl.Utf8,
            "Cls Cd": pl.Utf8,
            "Lf Cycle": pl.Utf8,
            "Color": pl.Utf8,
            "Finish": pl.Utf8,
            "BU": pl.Utf8,
            "BU Name": pl.Utf8,
            "M P G": pl.Utf8,
            "MPG Name": pl.Utf8,
            "IPG": pl.Utf8,
            "IPG Name": pl.Utf8,
            "Mkt Cls": pl.Utf8,
            "Mkt Cls Desc": pl.Utf8,
            "Prd Mgr": pl.Utf8,
            "Top Cust Ct": pl.Utf8,
            "Top Custs": pl.Utf8,
            "Sft Inv Method": pl.Utf8,
            "ABC Cd": pl.Utf8,
            "ABC Flg": pl.Utf8,
            "Sft Tm": pl.Float32,
            "Lead Tm Wks": pl.Int32,
            "Fcst Flg": pl.Utf8,
            "Plt Pln Min": pl.Int32,
            "Plt Lot Size": pl.Int32,
            "Std Pk": pl.Int32,
            "Inn Pk": pl.Utf8,
            "UPC": pl.Utf8,
            "Cart": pl.Utf8,
            "Pal Qty": pl.Int32,
            "Pal %": pl.Float32,
            "Cube": pl.Float32,
            "Length": pl.Float32,
            "Width": pl.Float32,
            "Height": pl.Float32,
            "Weight": pl.Float32,
            "On Hand": pl.Int32,
            "In Transit": pl.Int32,
            "QA Hold": pl.Int32,
            "Cust Ords": pl.Int32,
            "Net Inv": pl.Int32,
            "Tgt": pl.Int32,
            "Avg Wkly Dmd (6Mth)": pl.Float32,
            "Avg Wkly Dmd (2Mth)": pl.Float32,
            "Avg Wkly Fcst (2Mth)": pl.Float32,
            "Avg Wkly Fcst (6Mth)": pl.Float32,
            "Sft Time Qty": pl.Int32,
            "Base SS Qty": pl.Int32,
            "Sft Stk Qty": pl.Int32,
            "Sanning Indicator": pl.Utf8,
            "Excess Inv": pl.Int32,
            "Flag": pl.Utf8,
            "Inv Short": pl.Int32,
            "Flag2": pl.Utf8,
            "Wks of Supp Tgt (St)": pl.Float32,
            "Wks of Supp Act (St)": pl.Float32,
            "ST Var": pl.Float32,
            "Wks of Supp Tgt (Lt)": pl.Float32,
            "Wks of Supp Act (Lt)": pl.Float32,
            "LT Var": pl.Float32,
            "Dmd / Fcst": pl.Float32,
            "Notes": pl.Utf8,
            "EAU Units": pl.Int32,
            "WOS": pl.Float32,
            "Basic": pl.Utf8,
            "Dash": pl.Utf8,
            "San": pl.Utf8,
            "Basic-Dash": pl.Utf8,
        },
        "fg_status.parquet",
        True,
        None,
        "Sheet1",
        1
    ),
]

# Find all files based on prefix (case-insensitive)
def find_files(prefix: str, is_excel: bool) -> List[str]:
    ext = '.xlsx' if is_excel else '.csv'
    matching_files = [
        os.path.join(BASE_DIR, file)
        for file in os.listdir(BASE_DIR)
        if file.lower().startswith(prefix.lower()) and file.endswith(ext)
    ]
    if not matching_files:
        return []
    return sorted(matching_files)

# Convert Excel sheet to temporary CSV
def excel_to_csv(excel_file: str, sheet_name: str = "Sheet1", start_row: int = 1) -> str:
    wb = openpyxl.load_workbook(excel_file)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet {sheet_name} does not exist in {excel_file}. Available sheets: {wb.sheetnames}")
    sheet = wb[sheet_name]
    headers = [cell.value for cell in sheet[start_row] if cell.value is not None]
    data = []
    for row in sheet.iter_rows(min_row=start_row + 1):
        row_data = [cell.value if cell.value is not None else '' for cell in row]
        if len(row_data) > len(headers):
            row_data = row_data[:len(headers)]
        data.append(row_data)
    df = pd.DataFrame(data, columns=headers)
    temp_csv = os.path.join(BASE_DIR, f"temp_{os.path.basename(excel_file).replace('.xlsx', '.csv')}")
    df.to_csv(temp_csv, index=False)
    return temp_csv

# Memory and CPU tracking functions
process = psutil.Process()

def get_memory_usage() -> float:
    return process.memory_info().rss / (1024 * 1024)  # in MB

def get_cpu_times():
    return process.cpu_times()

# Extract unique values from input file for search columns
def get_search_values(df_lazy: pl.LazyFrame, search_columns: List[str]) -> Dict[str, list]:
    search_conditions = {}
    if not search_columns:
        return search_conditions
    
    df = df_lazy.select(search_columns).unique().collect()
    for col in search_columns:
        values = df[col].drop_nulls().to_list()
        if values:
            search_conditions[col] = values
    return search_conditions

# Process Parquet file: backup, search, remove, append
def process_parquet(
    input_path: str,
    parquet_file: str,
    dtype_dict: Dict[str, DataType],
    is_excel: bool,
    search_columns: Optional[List[str]],
    external_search_conditions: Optional[Dict[str, list]] = None,
    sheet_name: str = "Sheet1",
    start_row: int = 1
) -> Dict:
    memory_before = get_memory_usage()
    cpu_times_before = get_cpu_times()
    errors_encountered = []
    
    # Backup existing Parquet file if it exists
    backup_path = None
    if os.path.exists(parquet_file):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{os.path.basename(parquet_file).replace('.parquet', '')}_{timestamp}.parquet"
        backup_path = os.path.join(PARQUET_VERSIONS_DIR, backup_filename)
        try:
            shutil.copy2(parquet_file, backup_path)
            print(f"Backed up {parquet_file} to {backup_path}")
        except Exception as e:
            print(f"Warning: Failed to create backup for {parquet_file} at {backup_path}: {str(e)}")
            errors_encountered.append(str(e))
    
    # Read input data
    if is_excel:
        temp_csv = excel_to_csv(input_path, sheet_name=sheet_name, start_row=start_row)
        new_df_lazy = pl.scan_csv(temp_csv, schema_overrides=dtype_dict, ignore_errors=True)
    else:
        new_df_lazy = pl.scan_csv(input_path, schema_overrides=dtype_dict, ignore_errors=True)
    
    # Special handling for online_website_anaylsis
    if os.path.basename(parquet_file) == "online_website_anaylsis.parquet":
        # Rename column to match existing Parquet
        new_df_lazy = new_df_lazy.rename({"online product interaction conversion rate ": "online product interaction conversion rate"})
        new_df_lazy = new_df_lazy.with_columns([
            pl.col("online product interaction conversion rate").cast(pl.Float32),
            (pl.col("online pip visits +") * pl.col("online pip conversion rate +")).cast(pl.Float32).alias("order count TY"),
            (pl.col("online pip visits ly +") * pl.col("online pip conversion rate ly +")).cast(pl.Float32).alias("order count LY")
        ])
        new_df = new_df_lazy.collect()
        print(f"Processed online_website_anaylsis with {new_df.height} rows, added order count TY and LY")
        with open(log_file, 'a') as log:
            log.write(f"Processed online_website_anaylsis with {new_df.height} rows, added order count TY and LY at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
    
    # Special handling for online_classification
    elif os.path.basename(parquet_file) == "online_classification.parquet":
        new_df_lazy = new_df_lazy.with_columns(pl.col("online upc +").cast(pl.Utf8))
        new_df_lazy = new_df_lazy.filter(pl.col("online upc +").str.len_chars() == 12)
        new_df_lazy = new_df_lazy.group_by("oms id +").agg(pl.all().first())
        new_df = new_df_lazy.collect()
        unique_oms_id_count = new_df["oms id +"].n_unique()
        print(f"Processed online_classification with {new_df.height} rows, unique oms id + count: {unique_oms_id_count}")
        with open(log_file, 'a') as log:
            log.write(f"Processed online_classification with {new_df.height} rows, unique oms id + count: {unique_oms_id_count} at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
    
    # Group by UPC for FG Status Report to have one row per UPC
    elif "FG Status Report" in os.path.basename(input_path):
        new_df_lazy = new_df_lazy.with_columns(pl.concat_str([pl.col("Basic"), pl.lit("-"), pl.col("Dash")]).alias("Basic-Dash"))
        new_df_lazy = new_df_lazy.group_by("UPC").agg(pl.all().first())
        new_df = new_df_lazy.collect()
        unique_upc_count = new_df["UPC"].n_unique()
        print(f"Processed FG Status Report with {new_df.height} rows (grouped by UPC, unique UPC count: {unique_upc_count})")
        with open(log_file, 'a') as log:
            log.write(f"Processed FG Status Report with {new_df.height} rows (grouped by UPC, unique UPC count: {unique_upc_count}) at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
        if unique_upc_count != new_df.height or new_df.height < 15000 or new_df.height > 20000:
            print(f"Warning: Expected ~17822 rows for FG Status Report, got {new_df.height}")
            with open(log_file, 'a') as log:
                log.write(f"Warning: Expected ~17822 rows for FG Status Report, got {new_df.height} at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
    else:
        new_df = new_df_lazy.collect()
    
    new_row_count = new_df.height
    
    # Convert existing Parquet file schema for online_sales.parquet if necessary
    if os.path.basename(parquet_file) == "online_sales.parquet" and os.path.exists(parquet_file):
        existing_df_lazy = pl.scan_parquet(parquet_file)
        existing_schema = existing_df_lazy.collect_schema()
        if existing_schema.get("online return units +") == pl.Int32:
            print(f"Converting 'online return units +' to Float32 in {parquet_file}")
            temp_parquet = os.path.join(DATA_TABLES_DIR, f"temp_convert_{os.path.basename(parquet_file)}")
            existing_df_lazy.with_columns(
                pl.col("online return units +").cast(pl.Float32)
            ).sink_parquet(temp_parquet, compression="snappy")
            os.replace(temp_parquet, parquet_file)
    
    # Get search conditions from input file or use external conditions
    search_conditions = external_search_conditions if external_search_conditions else get_search_values(new_df_lazy, search_columns)
    
    rows_removed = 0
    final_row_count = new_row_count
    if search_columns and search_conditions and os.path.exists(parquet_file):
        existing_df_lazy = pl.scan_parquet(parquet_file)
        existing_row_count = existing_df_lazy.select(pl.len()).collect().item()
        
        # Validate search conditions
        for col in search_conditions.keys():
            if col not in search_columns:
                raise ValueError(f"Invalid search column '{col}' for {parquet_file}. Allowed: {search_columns}")
        
        # Corrected filter: keep rows unless BOTH day and oms id + match
        if "day" in search_conditions and "oms id +" in search_conditions:
            day_values = search_conditions["day"]
            if external_search_conditions and isinstance(day_values[0], str):
                day_values = [pd.to_datetime(val) for val in day_values]
            filter_expr = ~(pl.col("day").is_in(day_values) & pl.col("oms id +").is_in(search_conditions["oms id +"]))
        else:
            filter_expr = pl.lit(True)
        
        filtered_df_lazy = existing_df_lazy.filter(filter_expr)
        filtered_row_count = filtered_df_lazy.select(pl.len()).collect().item()
        rows_removed = existing_row_count - filtered_row_count
        
        # Collect new data and validate schema
        if new_df.schema != existing_df_lazy.collect_schema():
            print(f"Schema mismatch detected for {parquet_file}. Attempting to cast to match existing schema.")
            new_df = new_df.with_columns([
                pl.col(col).cast(dtype) for col, dtype in existing_df_lazy.collect_schema().items()
                if col in new_df.columns
            ])
        
        # Combine filtered existing data with new data
        combined_df_lazy = pl.concat([filtered_df_lazy, new_df.lazy()], how="vertical_relaxed")
        temp_parquet = os.path.join(DATA_TABLES_DIR, f"temp_{os.path.basename(parquet_file)}")
        combined_df_lazy.sink_parquet(temp_parquet, compression="snappy")
        os.replace(temp_parquet, parquet_file)
        final_row_count = filtered_row_count + new_row_count
    else:
        # Overwrite mode for no search_columns, no search_conditions, or no existing file
        new_df_lazy.sink_parquet(parquet_file, compression="snappy")
    
    # Clean up temporary CSV if created
    if is_excel and 'temp_csv' in locals() and os.path.exists(temp_csv):
        os.remove(temp_csv)
    
    memory_after = get_memory_usage()
    cpu_times_after = get_cpu_times()
    
    return {
        "memory_used": memory_after - memory_before,
        "user_time": cpu_times_after.user - cpu_times_before.user,
        "system_time": cpu_times_after.system - cpu_times_before.system,
        "search_conditions": search_conditions,
        "rows_removed": rows_removed,
        "rows_added": new_row_count,
        "final_row_count": final_row_count,
        "backup_path": backup_path if backup_path else "No backup",
        "errors": errors_encountered
    }

# Main processing
def main(external_search_conditions: Optional[Dict[str, list]] = None):
    start_time = time.time()
    summary = []
    
    for i, (prefix, dtype_dict, parquet_name, is_excel, search_columns, sheet_name, start_row) in enumerate(file_configs, 1):
        print(f"[{i}/{len(file_configs)}] Processing files starting with: {prefix}")
        file_paths = find_files(prefix, is_excel)
        if not file_paths and parquet_name != "online_classification.parquet":
            print(f"No {'Excel' if is_excel else 'CSV'} files found starting with {prefix}")
            summary.append({
                "file": f"No file found for prefix {prefix}",
                "parquet_file": parquet_name,
                "memory_used": 0.0,
                "user_time": 0.0,
                "system_time": 0.0,
                "search_conditions": {},
                "rows_removed": 0,
                "rows_added": 0,
                "final_row_count": 0,
                "backup_path": "No backup",
                "errors": [f"No {'Excel' if is_excel else 'CSV'} files found starting with {prefix}"]
            })
            continue
        
        parquet_file = os.path.join(DATA_TABLES_DIR, parquet_name)
        if file_paths:
            for j, file_path in enumerate(file_paths, 1):
                print(f"Processing file {j}/{len(file_paths)}: {file_path}")
                usage = process_parquet(file_path, parquet_file, dtype_dict, is_excel, search_columns, external_search_conditions, sheet_name, start_row)
                
                summary.append({
                    "file": file_path,
                    "parquet_file": parquet_name,
                    "memory_used": usage["memory_used"],
                    "user_time": usage["user_time"],
                    "system_time": usage["system_time"],
                    "search_conditions": usage["search_conditions"],
                    "rows_removed": usage["rows_removed"],
                    "rows_added": usage["rows_added"],
                    "final_row_count": usage["final_row_count"],
                    "backup_path": usage["backup_path"],
                    "errors": usage["errors"]
                })
                
                print(f"Finished processing {file_path} -> {parquet_name}")
    
    # Post-processing for FG Status Report merge into a single file
    fg_parquet = os.path.join(DATA_TABLES_DIR, "fg_status.parquet")
    classification_parquet = os.path.join(DATA_TABLES_DIR, "online_classification.parquet")
    merged_parquet = os.path.join(DATA_TABLES_DIR, "merged_classification.parquet")
    if os.path.exists(fg_parquet) and os.path.exists(classification_parquet):
        fg_df = pl.read_parquet(fg_parquet)
        classification_df = pl.read_parquet(classification_parquet)
        fg_df = fg_df.select(["UPC", "M P G", "MPG Name", "IPG", "IPG Name", "Prd Mgr", "Basic-Dash"]).with_columns(pl.col("UPC").cast(pl.Utf8))
        if "online upc +" in classification_df.columns:
            classification_df = classification_df.with_columns(pl.col("online upc +").str.slice(0, 11).cast(pl.Utf8).alias("UPC11"))
            merged_df = classification_df.join(
                fg_df,
                left_on="UPC11",
                right_on="UPC",
                how="left"
            ).drop("UPC11")
            # Ensure 1 row per oms id +
            duplicate_check = merged_df.group_by("oms id +").agg(pl.count().alias("count")).filter(pl.col("count") > 1)
            if duplicate_check.height > 0:
                print(f"Warning: {duplicate_check.height} duplicate oms id + found. Deduplicating by taking first values.")
                with open(log_file, 'a') as log:
                    log.write(f"Warning: {duplicate_check.height} duplicate oms id + found at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
                merged_df = merged_df.group_by("oms id +").agg(pl.all().first())
            else:
                print("Success: One row per unique oms id + in merged file")
                with open(log_file, 'a') as log:
                    log.write(f"Success: One row per unique oms id + in merged file at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
            merged_df.write_parquet(merged_parquet)
            print(f"Merged FG Status Report into online_classification.parquet and saved as {merged_parquet}")
            print(f"Final merged row count: {merged_df.height}, Unique oms id + count: {merged_df['oms id +'].n_unique()}")
            print(f"Merged fields from FG Status: M P G, MPG Name, IPG, IPG Name, Prd Mgr, Basic-Dash")
            with open(log_file, 'a') as log:
                log.write(f"Merged FG Status Report into {merged_parquet} with {merged_df.height} rows, unique oms id + count: {merged_df['oms id +'].n_unique()} at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
                log.write(f"Merged fields from FG Status: M P G, MPG Name, IPG, IPG Name, Prd Mgr, Basic-Dash\n")
        else:
            print("Skipping merge as 'online upc +' not found in online_classification.parquet")
            with open(log_file, 'a') as log:
                log.write(f"Skipping merge as 'online upc +' not found in online_classification.parquet at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
    else:
        if not os.path.exists(classification_parquet):
            print(f"Skipping merge as {classification_parquet} does not exist")
            with open(log_file, 'a') as log:
                log.write(f"Skipping merge as {classification_parquet} does not exist at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
        if not os.path.exists(fg_parquet):
            print(f"Skipping merge as {fg_parquet} does not exist")
            with open(log_file, 'a') as log:
                log.write(f"Skipping merge as {fg_parquet} does not exist at {datetime.now().strftime('%H:%M:%S %Z on %m/%d/%Y')}\n")
    
    print("\n### Summary ###")
    for item in summary:
        print(f"**File:** {item['file']} -> {item['parquet_file']}")
        print(f"  Backup: {item['backup_path']}")
        if item["errors"]:
            print(f"  Errors: {', '.join(item['errors'])}")
        if item["search_conditions"]:
            print(f"  Search conditions (truncated):")
            for col, values in item["search_conditions"].items():
                display_values = values[:5] if len(values) > 5 else values
                print(f"    {col}: {display_values}{'...' if len(values) > 5 else ''}")
            print(f"  Matching rows removed: {item['rows_removed']}")
            print(f"  New rows appended: {item['rows_added']}")
            print(f"  Final rows: {item['final_row_count']}")
        else:
            print(f"  Overwritten: {item['rows_added']} rows written")
            print(f"  Final rows: {item['final_row_count']}")
        print(f"  Memory used: {item['memory_used']:.2f} MB")
        print(f"  CPU user time: {item['user_time']:.2f} s")
        print(f"  CPU system time: {item['system_time']:.2f} s")
    print(f"\n**Merged Classification File Status**")
    if os.path.exists(merged_parquet):
        print(f"  {merged_parquet} created successfully")
    else:
        print(f"  {merged_parquet} not created")
    end_time = time.time()
    total_time = end_time - start_time
    print(f"\n**Total time taken:** {total_time:.2f} seconds")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert CSV/Excel to Parquet with search/remove/append")
    parser.add_argument("--search_conditions", type=str, help="JSON string of search conditions")
    args = parser.parse_args()
    
    external_search_conditions = None
    if args.search_conditions:
        external_search_conditions = json.loads(args.search_conditions)
    
    main(external_search_conditions)
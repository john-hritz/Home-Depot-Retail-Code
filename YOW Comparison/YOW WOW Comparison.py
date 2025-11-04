import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client as win32  # Requires: pip install pywin32

# --- CONFIGURATION ---
BASE_DIR = os.path.join(os.environ['USERPROFILE'], 'OneDrive - Leviton', 'Documents', 'Python', 'YOW Comparison')
email_to = "john.hritz@leviton.com"

# --- Load Excel Files ---
excel_files = [f for f in os.listdir(BASE_DIR) if f.endswith('.xlsx') and "Summary" not in f]
excel_files.sort(key=lambda x: int(''.join(filter(str.isdigit, x))))

file_low = os.path.join(BASE_DIR, excel_files[0])
file_high = os.path.join(BASE_DIR, excel_files[1])

df_low = pd.read_excel(file_low, engine='openpyxl')
df_high = pd.read_excel(file_high, engine='openpyxl')

df_low.columns = df_low.columns.str.strip()
df_high.columns = df_high.columns.str.strip()

# --- Convert OMSID to int and filter out zeros ---
df_low['OMSID'] = pd.to_numeric(df_low['OMSID'], errors='coerce').fillna(0).astype(int)
df_high['OMSID'] = pd.to_numeric(df_high['OMSID'], errors='coerce').fillna(0).astype(int)
df_low = df_low[df_low['OMSID'] != 0]
df_high = df_high[df_high['OMSID'] != 0]

# --- Normalize DC to first 4 characters ---
df_low['DC'] = df_low['DC'].astype(str).str[:4]
df_high['DC'] = df_high['DC'].astype(str).str[:4]

# --- Merge and Compare ---
key_cols = ['OMSID', 'DC']
merged = pd.merge(df_low, df_high, on=key_cols, how='outer', suffixes=('_Low', '_High'), indicator=True)

changes_rows = []
missing_in_higher_fw_omsids = set()
low_grouped = df_low.groupby(key_cols)
high_grouped = df_high.groupby(key_cols)

# --- NEW LOGIC: Identify OMSIDs completely missing in higher FW ---
omsids_in_low = set(df_low['OMSID'].unique())
omsids_in_high = set(df_high['OMSID'].unique())

omsids_missing_in_high = omsids_in_low - omsids_in_high

for omsid in omsids_missing_in_high:
    subset = df_low[df_low['OMSID'] == omsid]
    for rec in subset.to_dict('records'):
        rec['Change Type'] = 'OMSID Present LW Not Present TW'
        changes_rows.append(rec)
    missing_in_higher_fw_omsids.add(omsid)

# --- Continue with other comparisons ---
for _, row in merged.iterrows():
    key = tuple(row[k] for k in key_cols)
    if key[0] == 0 or key[0] in omsids_missing_in_high:
        continue  # Skip already handled OMSIDs

    if row['_merge'] == 'right_only':
        if key in high_grouped.groups:
            for rec in df_high.loc[high_grouped.groups[key]].to_dict('records'):
                rec['Change Type'] = 'New OMSID/DC Combo'
                changes_rows.append(rec)

    else:
        for col in ['Ntwk Replen', 'DC Replen']:
            col_low = f"{col}_Low"
            col_high = f"{col}_High"
            if col_low in row and col_high in row and row[col_low] == 'ON' and row[col_high] == 'OFF':
                if key in high_grouped.groups:
                    for rec in df_high.loc[high_grouped.groups[key]].to_dict('records'):
                        rec['Change Type'] = 'Replenishment Change OMSID/DC'
                        changes_rows.append(rec)
                break

# --- All DCs Off Inventory Now at 0 ---
inventory_cols = ['Avail OH Qty', 'OH Qty', 'CurrATSOHQty', 'OO Qty', 'Eff Inv']
valid_inventory_cols = [col for col in inventory_cols if col in df_low.columns and col in df_high.columns]

zero_inventory_omsids_list = []

omsids = df_low['OMSID'].unique()
for omsid in omsids:
    if omsid == 0:
        continue

    old_subset = df_low[df_low['OMSID'] == omsid]
    new_subset = df_high[df_high['OMSID'] == omsid]

    if old_subset.empty or new_subset.empty:
        continue

    if 'Ntwk Replen' not in df_low.columns or 'DC Replen' not in df_low.columns:
        continue

    old_replen_off = all(old_subset['Ntwk Replen'] == 'OFF') and all(old_subset['DC Replen'] == 'OFF')
    new_replen_off = all(new_subset['Ntwk Replen'] == 'OFF') and all(new_subset['DC Replen'] == 'OFF')

    if valid_inventory_cols:
        old_sum = old_subset[valid_inventory_cols].fillna(0).infer_objects(copy=False).sum().sum()
        new_sum = new_subset[valid_inventory_cols].fillna(0).infer_objects(copy=False).sum().sum()

        if old_replen_off and new_replen_off and old_sum > 0 and new_sum == 0:
            zero_inventory_omsids_list.append(omsid)
            for rec in new_subset.to_dict('records'):
                rec['Change Type'] = 'All DCs Off Inventory Now at 0'
                changes_rows.append(rec)

df_changes = pd.DataFrame(changes_rows)

# --- Create Summary Workbook ---
wb = Workbook()
ws1 = wb.active
ws1.title = os.path.basename(file_low).replace(".xlsx", "")
for r in dataframe_to_rows(df_low, index=False, header=True):
    ws1.append(r)
ws1.auto_filter.ref = ws1.dimensions

ws2 = wb.create_sheet(title=os.path.basename(file_high).replace(".xlsx", ""))
for r in dataframe_to_rows(df_high, index=False, header=True):
    ws2.append(r)
ws2.auto_filter.ref = ws2.dimensions

ws3 = wb.create_sheet(title="Changes")
if not df_changes.empty:
    for r in dataframe_to_rows(df_changes, index=False, header=True):
        ws3.append(r)
else:
    ws3.append(["No changes found."])
ws3.auto_filter.ref = ws3.dimensions
wb.active = wb.sheetnames.index("Changes")

summary_file = os.path.join(BASE_DIR, "Leviton_Comparison_Summary.xlsx")
wb.save(summary_file)

# --- Compose Email Body ---
change_counts = df_changes['Change Type'].value_counts()
email_body = "Leviton Report Comparison Summary:\n\n"
for label in ['New OMSID/DC Combo', 'Replenishment Change OMSID/DC', 'All DCs Off Inventory Now at 0', 'OMSID Present LW Not Present TW']:
    count = change_counts.get(label, 0)
    email_body += f"- {label}: {count}\n"

if zero_inventory_omsids_list:
    zero_inventory_omsids_str = ', '.join(map(str, sorted(set(zero_inventory_omsids_list))))
    email_body += f"\nAll DCs Off Inventory Now at 0: {zero_inventory_omsids_str}\n"

if missing_in_higher_fw_omsids:
    missing_fw_str = ', '.join(map(str, sorted(missing_in_higher_fw_omsids)))
    email_body += f"\nMissing in Higher FW OMSIDs: {missing_fw_str}\n"

email_body += "\nAttached is the comparison summary between the two Leviton reports. See the 'Changes' tab for details."

# --- Send Email via Outlook ---
outlook = win32.Dispatch('outlook.application')
mail = outlook.CreateItem(0)
mail.To = email_to
mail.Subject = "Leviton YOW TW vs LW Comparison Summary"
mail.Body = email_body
mail.Attachments.Add(summary_file)
mail.Send()

print("✅ Comparison complete and email sent via Outlook.")